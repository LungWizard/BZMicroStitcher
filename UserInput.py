import openpyxl
from ProcessData import Create_OMETIFF, Stitch

#USER INPUT
class UserInput():
    ImgDate = '03182021'
    OME_XLSX = 'OME-TIFF.xlsx'
    MainDirectory = 'C:/' + ImgDate + '/'
    WSI_List = openpyxl.load_workbook(MainDirectory+OME_XLSX).worksheets[0]
    SlideID = []
    GroupID = []
    XYID = []
    GCI_List = []
    ImgPath_List = []
    Path_List = []
    Num_Slides = WSI_List.max_row
    ErrorLog = open(MainDirectory + 'Error_Log_' + ImgDate + '.txt', 'a+')

    for X in range (1, Num_Slides):
        SlideID_Val = WSI_List.cell(row = X+1, column = 1).value
        SlideID.append(SlideID_Val)
        GroupID_Val = WSI_List.cell(row = X+1, column = 2).value
        GroupID.append(GroupID_Val)
        XYID_Val = WSI_List.cell(row = X+1, column = 3).value
        XYID.append(XYID_Val)
        
    for Y in range (0, Num_Slides-1):
        GCI_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' +
                        str(XYID[Y]) + '/' + '*.gci')
        
        ImgPath_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' +
                            str(XYID[Y]) + '/' + str(XYID[Y]) + '.tif') 
        
        Path_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' + 
                            str(XYID[Y]))


    for X in range (0, Num_Slides-1):
        try:
            Stitch(GCI_List[X], Path_List[X])
        except:
            ErrorLog.write("Something is not right, unable to stitch: " + GCI_List[X] + ", " + 
                  ImgPath_List[X] + ", " + SlideID [X] + "\n")
            pass
    
    for X in range (0, Num_Slides-1):
        try:
            Create_OMETIFF(MainDirectory, GCI_List[X], ImgPath_List[X], SlideID[X])
        except:
            ErrorLog.write("Something is not right, unable to create OME-TIFF: " 
                           + GCI_List[X] + ", " + ImgPath_List[X] + ", " + SlideID [X] + "\n")
            pass

    ErrorLog.write("If this is the only line you see then everything went well! YAY!")
    ErrorLog.close()